#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :  excel.py
@Time    :  2024/04/21 21:03:43
@Version :  1.0
@Desc    :  表格模块
'''
import os
from typing import Union
import traceback
import pandas as pd
from app.utils.logger import crmLogger
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

def readExcel(filePath: str) -> Union[pd.DataFrame, None]:
    '''
    读取表格
    :param filepath: 文件路径
    :return:
    '''
    try:
        return pd.read_excel(filePath).fillna(value="")  # bugfix:替换表中的空值nan
    except:
        crmLogger.error(f"[readExcel]读取表格时发生错误: {traceback.format_exc()}")
        return None


def createExcel(filepath: str, filename: str, sheet_name: str, header: dict, data: dict, styles: dict, is_template: bool = False, passwd: str="") -> bool:
    '''
    创建表格
    :param filepath: 文件存放目录
    :param filename: 文件名
    :param sheet_name: 工作簿名称
    :param header: 表头 {"column_name": {"index": "A/B/C...", "must_input": True/False}}
    :param data: 数据 {"clumn_name": ["d1", "d2"...]}
    :param styles: 样式 {"column_name": {"index": "A/B/C...", "options": "选项1,选项2..."}}"}
    :param is_template: 是否为模板
    :param passwd: 加密密码
    :return:
    '''
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name  # 设置工作簿名称
        # 写入表头
        for name, v in header.items():
            if v["must_input"]:  # 必填选项
                red = InlineFont(color="FF0000")  # 设置字体颜色
                richText = TextBlock(red, "*")
                ws[f"{v['index']}1"] = CellRichText([name, richText])
            else:
                ws[f"{v['index']}1"] = name
            ws.column_dimensions[f"{v['index']}"].width = len(name) * 4  # 设置列宽度
        # 写入数据
        if not is_template:
            for col, row in data.items():
                col_index = header[col]["index"]
                for idx, val in enumerate(row, start=2):  # 从第二行开始写入数据
                    ws[f"{col_index}{idx}"] = val
        # 设置下拉样式
        if styles:
            for _, val in styles.items():
                dv = DataValidation(type="list", formula1=f'"{val["options"]}"', showDropDown=False, allow_blank=True)  # showDropDown-是否显示下拉箭头,allow_blank-是否允许空值
                ws.add_data_validation(dv)
                dv.add(f"{val['index']}2:{val['index']}{1000 if ws.max_row < 1000 else ws.max_row}")
        # 保护工作表保护不被修改
        # ws.protection.sheet = True
        # ws.protection.enable()
        # 保存文件
        wb.save(os.path.join(filepath, f"{filename}.xlsx"))
        # 给表格添加密码
        # 仅支持在windows系统下使用
        if passwd and os.name == "nt":
            import win32com.client
            xcl = win32com.client.Dispatch("Excel.Application")
            xwb = xcl.Workbooks.Open(os.path.join(filepath, f"{filename}.xlsx"), False, False, None, "")
            xcl.DisplayAlerts = False  # 关闭告警对话框
            xwb.SaveAs(os.path.join(filepath, f"{filename}.xlsx"), None, passwd, "")  # 设置密码
            xcl.Quit()
        return True
    except:
        crmLogger.error(f"[createExcel]写入表格时发生错误: {traceback.format_exc()}")
        return False
