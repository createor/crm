#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :  task.py
@Time    :  2024/06/29 22:20:49
@Version :  1.0
@Desc    :  任务模块
'''
import os
import json
import traceback
import re
import queue
import threading
from datetime import datetime, date, timedelta
from app.utils import TEMP_DIR, UPLOAD_EXCEL_DIR, redisClient, crmLogger, createExcel, readExcel, scan_ip, getUuid
from app.src.models import engine, db_session, Task, DetectResult, Notice, History, Header, Options, File, MyHeader, initManageTable
from sqlalchemy import and_, insert, update
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils import get_column_letter
import pandas as pd

importExecutor = ThreadPoolExecutor(max_workers=5)  # 导入线程池
exportExecutor = ThreadPoolExecutor(max_workers=5)  # 导出线程池
pingExecutor = ThreadPoolExecutor(max_workers=5)    # ping线程池
ip_queue = queue.Queue()     # ip队列
result_queue = queue.Queue() # 结果队列

def notifyTask(task_id: str, name: str, table_id: str, table_name: str, keyword: str, before: int=0) -> None:
    '''
    通知任务
    :param task_id: 任务id
    :param name: 表名
    :param table_id: 表id
    :param table_name: 表别名
    :param keyword: 日期字段
    :param before: 提前天数
    '''
    manageTable = initManageTable(table_name)        # 实例化已存在资产表

    try:      # 查询过期数据数量
        if before == 0:
            today = datetime.now().strftime("%Y-%m-%d")  # 获取今天日期
        else:
            today = (date.today() + timedelta(days=before)).strftime("%Y-%m-%d")
        expire_data = db_session.query(manageTable).filter(and_(f"{today} 00:00:00" <= getattr(manageTable.c, keyword), getattr(manageTable.c, keyword) <= f"{today} 23:59:59")).count()
    finally:
        db_session.close()
        
    if expire_data > 0:
        try:  # 写入数据库
            if before == 0:
                _msg = f"你的资产表{name}有{expire_data}条数据今天就要过期了"
            else:
                _msg = f"你的资产表{name}有{expire_data}条数据还有{before}天就要过期了"
            notice_data = Notice(message=_msg, notify_id=task_id, href=f"id={table_id}&key={keyword}&value={today}")
            db_session.add(notice_data)
            db_session.commit() 
        except:
            db_session.rollback()
            crmLogger.error(f"[notifyTask]写入notice表发生异常: {traceback.format_exc()}")
        finally:
            db_session.close()
    crmLogger.info(f"[notifyTask]定时任务执行成功: 任务id({task_id})")

def writeError(task_id: str, error: str, status: int=3) -> None:
    '''
    导入任务写入错误信息
    :param task_id: 任务id
    :param error: 错误信息
    :param status: 状态,默认值为3
    '''
    filename = getUuid()
    filepath = os.path.join(TEMP_DIR, f"{filename}.txt")

    with open(filepath, "w", encoding="utf-8") as f:  # 创建错误文件
        f.write(error)

    try:  # 写入file表
        err_file = File(uuid=filename, filename="错误信息.txt", filepath=0, affix="txt")
        db_session.add(err_file)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[writeError]写入file表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()
        if status == 3:
            redisClient.setData(f"crm:task:{task_id}", json.dumps({"error": f"{error}", "speed": 100}), 300)

    try:  # 更新history表状态和错误文件
        db_session.query(History).filter(History.id == task_id).update({"status": status, "err_file": filename})
        db_session.commit()
    except:     
        db_session.rollback()
        crmLogger.error(f"[writeError]更新history表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()
        if status == 3:
            redisClient.setData(f"crm:task:{task_id}", json.dumps({"error": f"{error}", "speed": 100}), 300)

def importTableTask(table_name: str) -> None:
    '''
    表格导入任务
    :param table_name: 表别名
    '''
    lock = redisClient.lock(f"import_{table_name}_lock", timeout=300)  # 设置300秒锁过期时间

    if lock.acquire(blocking=False): # 获取锁

        try:
            while redisClient.llen(f"crm:import:{table_name}") > 0:    # 导入队列不为空时

                is_continue = False  # 是否继续执行循环

                task_data = json.loads(redisClient.rpop(f"crm:import:{table_name}"))  # 读取任务数据

                try:                 # 更新任务状态为执行中
                    db_session.query(History).filter(History.id == task_data["task_id"]).update({"status": 1})
                    db_session.commit()
                except:
                    db_session.rollback()
                    crmLogger.error(f"[importTableTask]更新history表发生异常: {traceback.format_exc()}")
                    continue
                finally:
                    db_session.close()

                crmLogger.debug(f"[importTableTask]正在执行导入任务: {task_data}")

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 5}), 300)  # 设置进度为5%
                
                temp_table = readExcel(os.path.join(UPLOAD_EXCEL_DIR, task_data["file"]))  # 读取表格

                if temp_table is None:  # 读取表格失败
                    crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: 读取导入表格失败")
                    writeError(task_data["task_id"], "读取导入表格失败")
                    continue

                table_headers = temp_table.columns.tolist()  # 获取表头字段

                if len(table_headers) == 0:                  # 如果表头为空
                    crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: 读取表格表头为空")
                    writeError(task_data["task_id"], "读取表格表头为空")
                    continue

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 15}), 300)  # 设置进度为15%

                templ_header = redisClient.getData(f"crm:header:{task_data['table']}")  # 从redis中读取缓存

                if templ_header:
                    templ_header = [MyHeader(i) for i in json.loads(templ_header)]
                else:
                    try:  # 查询数据中表头的记录
                        templ_header = db_session.query(Header.name, Header.value, Header.value_type, Header.type, Header.must_input, Header.is_unique).filter(Header.table_name == task_data["table"]).all()
                    finally:
                        db_session.close()

                if task_data["update"]:
                    _t_h = [h.name for h in templ_header]
                    _t_h.append("_id")
                    if _t_h.sort() != list(map(lambda x: x.rsplit("*", 1)[0] if x.endswith("*") else x, table_headers)).sort():  # 去除*号后排序后比较
                        crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: 缺失字段")
                        writeError(task_data["task_id"], "缺失字段")
                        continue
                else:
                    # 读取文件的header和数据库中比对是否一致,不一致说明不是使用模板导入
                    if [h.name for h in templ_header].sort() != list(map(lambda x: x.rsplit("*", 1)[0] if x.endswith("*") else x, table_headers)).sort():  # 去除*号后排序后比较
                        crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: 未使用模板导入")
                        writeError(task_data["task_id"], "未使用模板导入")
                        continue

                manageTable = initManageTable(task_data["table"])  # 实例化已存在的资产表

                # 判断必填值列是否有空值
                must_header = [h.name for h in templ_header if h.must_input == 1]
                # bugfix:值为全数字或者类型为数字导入报错
                def checkIsNull(x):
                    if x == "":
                        return True
                    else:
                        return False
                for h in must_header:
                    # if temp_table[f"{h}*"].isnull().any():         # 判断是否有空值,带*表示必填
                    if temp_table[f"{h}*"].apply(checkIsNull).any():  # bugfix:pandas使用fillna填充后空值变为"""
                        crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: {h}字段为必填项,存在空值")
                        writeError(task_data["task_id"], f"{h}字段为必填项,存在空值")
                        is_continue = True
                        break
                
                if is_continue:
                    continue

                unique_header = []
                option_header = []
                length_header = []
                for h in templ_header:
                    if h.value_type == 2:
                        if h.must_input == 1:
                            length_header.append({"name": f"{h.name}*", "value": h.value, "length": h.length})
                        else:
                            length_header.append({"name": h.name, "value": h.value, "length": h.length})
                    if h.type == 2:
                        if h.must_input == 1:
                            option_header.append({"name": f"{h.name}*", "value": h.value})
                        else:
                            option_header.append({"name": h.name, "value": h.value})
                    if h.is_unique == 1:
                        if h.must_input == 1:
                            unique_header.append({"name": f"{h.name}*", "value": h.value})
                        else:
                            unique_header.append({"name": h.name, "value": h.value})

                # 校验数据是否唯一,有重复导入     
                for h in unique_header:
                    if temp_table[f"{h['name']}"].duplicated().any():  # 判断是否有重复数据
                        crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: {h['name'].rsplit('*', 1)[0]}字段为唯一项,有重复数据")
                        writeError(task_data["task_id"], f"{h['name'].rsplit('*', 1)[0]}字段为唯一项,有重复数据")
                        is_continue = True
                        break
                    
                    if task_data["update"]:
                        try:
                            col_data = db_session.query(getattr(manageTable.c, h["value"])).filter(manageTable.c._id.notin_(temp_table["_id"].tolist())).all()
                        finally:
                            db_session.close()
                    else:
                        try:  # 比对导入的表格数据与数据库中是否有重复
                            col_data = db_session.query(getattr(manageTable.c, h["value"])).all()
                        finally:
                            db_session.close()

                    col_data = [getattr(c, h["value"]) for c in col_data if getattr(c, h["value"])]

                    if temp_table[f"{h['name']}"].isin(col_data).any():  # 判断是否有重复数据
                        crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: {h['name'].rsplit('*', 1)[0]}字段在数据库中有重复数据")
                        writeError(task_data["task_id"], f"{h['name'].rsplit('*', 1)[0]}字段在数据库中有重复数据")
                        is_continue = True
                        break
                
                if is_continue:
                    continue

                # 校验数据是否从下拉列表选项中值
                for h in option_header:        
                    try:
                        _opt = db_session.query(Options.option_name).filter(Options.table_name == task_data["table"], Options.header_value == h["value"]).all()
                    finally:
                        db_session.close()
                    if not temp_table[f"{h['name']}"].isin([o.option_name for o in _opt]).all():
                        crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: {h['name'].rsplit('*', 1)[0]}字段存在非固定选项值")
                        writeError(task_data["task_id"], f"{h['name'].rsplit('*', 1)[0]}字段存在非固定选项值")
                        is_continue = True
                        break

                if is_continue:
                    continue

                # 校验数据中长度是否合法
                for h in length_header:
                    # is_all_length = temp_table[h.name].str.len == h.length
                    # if not is_all_length.all():
                    if len(temp_table[temp_table[f"{h['name']}"].astype(str).str.len() != h["length"]]) > 0:  # bugfix:转换为文本类型再获取长度
                        crmLogger.error(f"[importTableTask]用户{task_data['user']}导入资产表{task_data['table']}失败: {h['name'].rsplit('*', 1)[0]}字段存在不满足长度的值")
                        writeError(task_data["task_id"], f"{h['name'].rsplit('*', 1)[0]}字段存在不满足长度的值")
                        is_continue = True
                        break
                
                if is_continue:
                    continue

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 50}), 300)  # 设置进度为50%
                                
                insertOrUpdate_data = temp_table.to_dict(orient="records")

                date_header = [h.name for h in templ_header if h.value_type == 4]
                datetime_header = [h.name for h in templ_header if h.value_type == 5]

                err_msg = []
                for idx, i in enumerate(insertOrUpdate_data, start=1):
                    if task_data["update"]:
                        i["_update_user"] = task_data["user"]
                    else:
                        i["_create_user"] = task_data["user"]
                    for c in templ_header:
                        if c.must_input == 1:
                            new_data = i.pop(f"{c.name}*")
                        else:
                            new_data = i.pop(c.name)
                        if new_data and not pd.isna(new_data):  # bugfix:pandas读取非法时间NaT
                            if c.name in date_header:       # 如果是时间,判断是否是date类型,否则进行转换
                                if isinstance(new_data, datetime):
                                    new_data = date.fromordinal(new_data.toordinal())
                                elif not isinstance(new_data, date):
                                    try:
                                        new_data = datetime.strptime(new_data, "%Y-%m-%d")
                                    except (ValueError, TypeError):
                                        crmLogger.error(f"[importTableTask]第{idx}行,{c.name}字段值{new_data}不符合日期格式")
                                        err_msg.append(f"第{idx}行,{c.name}字段值{new_data}不符合日期格式,请检查")
                                        continue
                            elif c.name in datetime_header:  # 如果是时间,判断是否是datetime类型,否则进行转换
                                if isinstance(new_data, date):
                                    new_data = datetime.fromordinal(new_data.toordinal())
                                elif not isinstance(new_data, datetime):
                                    try:
                                        new_data = datetime.strptime(new_data, "%Y-%m-%d %H:%M:%S")
                                    except (ValueError, TypeError):
                                        crmLogger.error(f"[importTableTask]第{idx}行,{c.name}字段值{new_data}不符合时间格式")
                                        err_msg.append(f"第{idx}行,{c.name}字段值{new_data}不符合时间格式,请检查")
                                        continue
                            else:
                                if isinstance(new_data, datetime):
                                    new_data = new_data.strftime("%Y-%m-%d %H:%M:%S")
                                elif isinstance(new_data, date):
                                    new_data = new_data.strftime("%Y-%m-%d")
                                else:
                                    new_data = str(new_data)
                            i.update({c.value: new_data})
                        else:
                            if c.name in date_header or c.name in datetime_header:
                                i.update({c.value: None})
                            else:
                                i.update({c.value: ""})

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 80}), 300) # 设置进度为80%

                try:
                    with engine.begin() as conn:  # 开启事务,批量插入或更新数据
                        if task_data["update"]:   # 更新数据
                            for record in insertOrUpdate_data:
                                _update_id = record.pop("_id")
                                stmt = update(manageTable).where(manageTable.c._id == _update_id).values(record)
                                conn.execute(stmt)
                        else:  # 插入数据
                            stmt = insert(manageTable)
                            conn.execute(stmt, insertOrUpdate_data)
                except:
                    crmLogger.error(f"[importTableTask]资产表{table_name}{'更新' if task_data['update'] else '插入'}数据发生异常: {traceback.format_exc()}")
                    continue

                if len(err_msg) > 0:
                    writeError(task_data["task_id"], "\n".join(err_msg), 2)
                else:
                    try:
                        db_session.query(History).filter(History.id == task_data["task_id"]).update({"status": 2})
                        db_session.commit()
                    except:
                        db_session.rollback()
                        crmLogger.error(f"[importTableTask]更新history表发生异常: {traceback.format_exc()}")
                        continue
                    finally:
                        db_session.close()

                crmLogger.info(f"[importTableTask]资产表{table_name}导入成功")
                
                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 100}), 300)   # 设置进度为100%

                # 导入成功,要删除图表缓存
                redisClient.delData(f"crm:echart:{task_data['table']}")

        except:
            crmLogger.error(f"[importTableTask]资产表{table_name}导入失败: {traceback.format_exc()}")

        finally:
            lock.release()  # 释放锁

    else:
        crmLogger.info("[importTableTask]资产表{table_name}导入任务已被其他进程占用")

def startImportTableTask(table_name: str):
    '''启动表格导入任务'''
    return importExecutor.submit(importTableTask, (table_name))

def exportTableTask(table_name: str):
    '''
    表格导出任务
    :param table_name: 表名
    '''
    lock = redisClient.lock(f"export_{table_name}_lock", timeout=300)  # 设置300秒锁过期时间

    if lock.acquire(blocking=False):  # 获取锁

        try:
            while redisClient.llen(f"crm:export:{table_name}") > 0:
                
                task_data = json.loads(redisClient.rpop(f"crm:export:{table_name}"))  # 获取任务数据

                try:  # 数据库中设置次任务状态
                    db_session.query(History).filter(History.id == task_data["task_id"]).update({"status": 1})
                    db_session.commit()
                except:
                    db_session.rollback()
                    crmLogger.error(f"[exportTableTask]更新history表发生异常: {traceback.format_exc()}")
                    continue
                finally:
                    db_session.close()

                crmLogger.debug(f"[exportTableTask]正在执行导出任务: {task_data}")
                
                # 将任务进度写入redis
                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 5}), 300)

                manageTable = initManageTable(task_data["table"])  # 实例化已存在的资产表

                if task_data["filter"]:  # 如果存在筛选条件
                    try:
                        filter = json.loads(task_data["filter"].replace("'",""))
                        if filter["type"] == 1:
                            if "c" in filter and filter["c"] and filter["c"] == "eq":
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]) == filter["value"]).all()
                            else:
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]).like(f"%{filter['value']}%")).all()
                        elif filter["type"] == 2:
                            export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]) == filter["value"]).all()
                        elif filter["type"] == 3:
                            if filter["c"] == "eq":
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]) == filter["value"]).all()
                            elif filter["c"] == "gt":
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]) > filter["value"]).all()
                            elif filter["c"] == "lt":
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]) < filter["value"]).all()
                            elif filter["c"] == "ge":
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]) >= filter["value"]).all()
                            elif filter["c"] == "le":
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]) <= filter["value"]).all()
                            elif filter["c"] == "ne":
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]) != filter["value"]).all()
                            elif filter["c"] == "bt":
                                export_data = db_session.query(manageTable).filter(getattr(manageTable.c, filter["key"]).between(f"{filter['value']} 00:00:00", f"{filter['value']} 23:59:59")).all()
                    finally:
                        db_session.close()
                else:  # 全量导出
                    try:
                        export_data = db_session.query(manageTable).all()
                    finally:
                        db_session.close()

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 15}), 300)

                header = redisClient.getData(f"crm:header:{task_data['table']}")

                if header:
                    header = [MyHeader(i) for i in json.loads(header)]
                else:
                    try:
                        header = db_session.query(Header.name, Header.type, Header.value, Header.value_type, Header.must_input, Header.order).filter(Header.table_name == task_data["table"]).order_by(Header.order.asc()).all()
                    finally:
                        db_session.close()

                table_header = {}
                table_styles = {}

                if task_data["update"]:
                    table_header["_id"] = {
                        "name": "_id",
                        "index": "A",
                        "must_input": False
                    }

                for h in header:    
                    table_header[h.name] = {
                        "name": h.value,
                        "index": get_column_letter(h.order + 2) if task_data["update"] else get_column_letter(h.order + 1),
                        "must_input": h.must_input == 1
                    }

                    if h.type == 2:  # 如果是下拉列表
                        try:
                            opt = db_session.query(Options.option_name).filter(Options.table_name == task_data["table"], Options.header_value == h.value).all()
                        finally:
                            db_session.close()

                        table_styles[f"{h.name}"] = {
                            "index": get_column_letter(h.order + 2) if task_data["update"] else get_column_letter(h.order + 1),
                            "options": ",".join([o.option_name for o in opt])
                        }

                write_data = {}  # 要写入表格的数据
                for h, v in table_header.items():
                    write_data[h] = [getattr(i, v["name"]) if getattr(i, v["name"]) else "" for i in export_data]

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 50}), 300)

                result_status = 2
                if not createExcel(filepath=TEMP_DIR, filename=task_data["task_id"], sheet_name=task_data["name"], header=table_header, data=write_data, styles=table_styles, passwd=task_data["password"]):
                    result_status = 3

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 80}), 300)

                try:
                    db_session.query(History).filter(History.id == task_data["task_id"]).update({"status": result_status, "file_uuid": f"{task_data['task_id'] if result_status == 2 else ''}"})
                    db_session.commit()
                except:
                    db_session.rollback()
                    crmLogger.error(f"[exportTableTask]更新history表发生异常: {traceback.format_exc()}")
                    redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "数据库异常", "speed": 100}), 300)
                    continue
                finally:
                    db_session.close()

                if result_status == 2:
                    try:
                        export_file = File(uuid=task_data["task_id"], filename=f"{task_data['name']}资产表导出文件.xlsx", filepath=0, upload_user=task_data["user"], password=f"{task_data['password'] if task_data['password'] else ''}", affix="xlsx")
                        db_session.add(export_file)
                        db_session.commit()
                    except:
                        db_session.rollback()
                        crmLogger.error(f"[exportTableTask]更新file表发生异常: {traceback.format_exc()}")
                        redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "数据库异常", "speed": 100}), 300)
                        continue
                    finally:
                        db_session.close()

                crmLogger.info(f"[exportTableTask]资产表{table_name}导出成功")

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 100}), 300)
        
        except:
            crmLogger.error(f"[exportTableTask]资产表{table_name}导出失败: {traceback.format_exc()}")

        finally:
            lock.release()
    else:
        crmLogger.info(f"[exportTableTask]资产表{table_name}导出任务已被其他进程占用")

def startExportTableTask(table_name: str):
    '''启动表格导出任务'''
    return exportExecutor.submit(exportTableTask, (table_name))

def consumer(ip_queue: queue.Queue, result_queue: queue.Queue):
    '''
    消费者
    :param ip_queue: ip队列
    :param result_queue: 结果队列
    '''
    while not ip_queue.empty():
        ip = ip_queue.get()
        if ip:
            status = int(scan_ip(ip))
            result_queue.put({
                "ip": ip,
                "status": status,
                "reason": ""
            })

def pingHostTask():
    '''批量探测主机任务'''
    lock = redisClient.lock("ping_lock", timeout=300)  # 300秒锁过期时间

    if lock.acquire(blocking=False):  # 获取锁

        try:
            while redisClient.llen("crm:task:ping") > 0:

                task_data = json.loads(redisClient.rpop("crm:task:ping"))

                try:
                    db_session.query(Task).filter(Task.id == task_data["task_id"]).update({"status": 1})
                    db_session.commit()
                except:
                    db_session.rollback()
                    crmLogger.error(f"[pingHostTask]更新task表发生异常: {traceback.format_exc()}")
                    continue
                finally:
                    db_session.close()

                crmLogger.debug(f"[pingHostTask]开始探测主机任务: {task_data}")

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 5}), 500)

                try:
                    manageTable = initManageTable(task_data["table"])
                    ip_result = db_session.query(getattr(manageTable.c, task_data["keyword"])).all()
                finally:
                    db_session.close()

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 15}), 500)

                if ip_result:
                    for i in ip_result:
                        # 使用正则校验是否是IP地址
                        if re.match(r"^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$", i[0]):
                            ip_queue.put(i[0])
                        else:
                            result_queue.put({
                                "ip": i[0],
                                "status": 2,
                                "reason": "不是一个正确的IP地址"
                            })
                else:
                    continue

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 30}), 500)

                threads = []
                for _ in range(5):
                    thread = threading.Thread(target=consumer, args=(ip_queue, result_queue,))
                    thread.start()
                    threads.append(thread)

                for t in threads:
                    t.join()

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 50}), 500)

                # 批量插入结果中
                insert_data = []
                while not result_queue.empty():
                    result = result_queue.get()
                    if result:
                        insert_data.append({
                            "ip": result["ip"],
                            "reason": result["reason"] if result["reason"] else "",
                            "status": result["status"],
                            "task_id": task_data["task_id"]
                        })

                if insert_data:
                    try:
                        with engine.connect() as conn:
                            stmt = insert(DetectResult)
                            conn.execute(stmt, insert_data)
                    except:
                        crmLogger.error(f"[pingHostTask]批量插入detect_result表发生异常: {traceback.format_exc()}")
                        redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "数据库异常", "speed": 100}), 500)
                        continue

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 80}), 500)

                try:
                    db_session.query(Task).filter(Task.id == task_data["task_id"]).update({"status": 2})
                    db_session.commit()
                except:
                    db_session.rollback()
                    crmLogger.error(f"[pingHostTask]更新task表发生异常: {traceback.format_exc()}")
                    redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "数据库异常", "speed": 100}), 500)
                    continue
                finally:
                    db_session.close()

                redisClient.setData(f"crm:task:{task_data['task_id']}", json.dumps({"error": "", "speed": 100}), 500)

                crmLogger.info(f"[pingHostTask]ping主机任务{task_data['task_id']}结束")   

        except:
            crmLogger.error(f"[pingHostTask]ping主机任务发生异常: {traceback.format_exc()}")

        finally:
            lock.release()
    else:
        crmLogger.warning(f"[pingHostTask]ping主机任务被其他进程占用")

def startPingTask():
    '''启动ping任务'''
    return pingExecutor.submit(pingHostTask)
