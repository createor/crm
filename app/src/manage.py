#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :  manage.py
@Time    :  2024/04/26 13:53:24
@Version :  1.0
@Desc    :  资产管理模块
'''
import os
import re
import time
import json
import traceback
from flask import Blueprint, Response, request, g, jsonify
from app.utils import UPLOAD_EXCEL_DIR, TEMP_DIR, SYSTEM_DEFAULT_TABLE, methods, crmLogger, redisClient, job, readExcel, createExcel, getUuid, verify, converWords, undesense, formatDate
from app.src.models import engine, db_session, Manage, Header, Log, Options, Echart, Task, DetectResult, File, Notify, History, MyHeader, initManageTable, generateManageTable, addColumn, alterColumn, refreshMeta
from app.src.task import startExportTableTask, startImportTableTask, startPingTask
from sqlalchemy import or_, func, Column, String
from sqlalchemy.sql import insert
from sqlalchemy.orm import mapper
from datetime import datetime, date
from openpyxl.utils import get_column_letter

manage = Blueprint("manage", __name__)

@manage.route("/query", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查询资产表", check_ip=True)
def queryTable():
    '''查询所有资产表'''
    args = request.args                        # 获取请求参数

    title = args.get("title", None)            # 资产表标题
    page = int(args.get("page", default=1))    # 当前页码,默认第一页
    limit = int(args.get("limit", default=3))  # 每页显示数量,默认3条

    if title:  # 如果存在标题搜索

        try:
            count = db_session.query(Manage).filter(Manage.name.like("%{}%".format(title))).count()
        finally:
            db_session.close()

        if count == 0:  # 如果没有搜索到结果,直接返回空列表
            return jsonify({"code": 0, "message": {"total": 0, "data": []}}), 200

        try:  # 模糊搜索结果,按最新创建的表格降序
            result = db_session.query(Manage).filter(Manage.name.like("%{}%".format(title))).order_by(Manage.create_time.desc()).offset((page - 1) * limit).limit(limit).all()  
        finally:
            db_session.close()

    else:  # 不存在标题搜索

        try:
            count = db_session.query(Manage).count()
        finally:
            db_session.close()

        if count == 0:
            return jsonify({"code": 0, "message": {"total": 0, "data": []}}), 200

        try:
            result = db_session.query(Manage).order_by(Manage.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
        finally:
            db_session.close()

    try:  # 写入log表
        query_log = Log(ip=g.ip_addr, operate_type="查询资产表", operate_content=f"查询资产表,title={title}", operate_user=g.username)
        db_session.add(query_log)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[queryTable]写入log表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()

    crmLogger.info(f"[queryTable]用户{g.username}成功查询了资产表,title={title},page={page},limit={limit}")  # 写入日志文件

    return jsonify({
        "code": 0,
        "message": {
            "total": count,
            "data": [{"id": item.uuid, "title": item.name, "remark": item.description, "image": f"/crm/api/v1/images/{item.table_image}", "time": f"{item.create_time}创建"} for item in result]
        }
    }), 200

@manage.route("/all", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查询所有资产表", check_ip=True)
def getAllTable():
    '''查询所有资产表'''
    try:
        result = db_session.query(Manage.name, Manage.uuid).all()
    finally:
        db_session.close()
    
    return jsonify({"code": 0, "message": [{"id": item.uuid, "title": item.name} for item in result]}), 200

@manage.route("/ip_col", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查询资产表IP列", check_ip=True)
def getTableIpCol():
    '''查询资产表IP列'''
    args = request.args                # 获取请求参数

    table_uuid = args.get("id", None)
    ip_col_name = args.get("ip_col", None)

    if not table_uuid or not ip_col_name:
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:
        table = db_session.query(Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:
        col_name = db_session.query(Header.value).filter(Header.table_name == table.table_name, Header.name == ip_col_name).first()
    finally:
        db_session.close()

    if not col_name:
        return jsonify({"code": -1, "message": "IP列不存在"}), 400
    
    # 删除缓存的header
    redisClient.delData(f"crm:header:{table.table_name}")
    
    return jsonify({"code": 0, "message": col_name.value}), 200

@manage.route("/bind", methods=methods.ALL)
@verify(allow_methods=["POST"], module_name="绑定资产表IP列", check_ip=True)
def bindTableIpCol():
    '''绑定资产表IP列'''
    reqData = request.get_json()

    if not all(key in reqData for key in ["table_id", "ip_col"]):  # 校验参数
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    table_uuid = reqData["table_id"]
    ip_column = reqData["ip_col"]

    if not all([table_uuid, ip_column]):
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):  # 校验资产表是否存在
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:
        return jsonify({"code": -1, "message": "资产表不存在"}), 400

    try:
        exist_col = db_session.query(Header.value).filter(Header.table_name == table.table_name, Header.is_ip == 1).first()
        if exist_col and exist_col.value != ip_column:
            exist_col.is_ip = 0
    except:
        db_session.rollback()
        crmLogger.error(f"[bindTableIpCol]更新header表发生异常: {traceback.format_exc()}")
        return jsonify({"code": -1, "message": "数据库异常"}), 500
    finally:
        db_session.close()

    try:
        db_session.query(Header).filter(Header.table_name == table.table_name, Header.value == ip_column).update({"is_ip": 1})
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[bindTableIpCol]更新header表发生异常: {traceback.format_exc()}")
        return jsonify({"code": -1, "message": "数据库异常"}), 500
    finally:
        db_session.close()
     
    crmLogger.info(f"[bindTableIpCol]用户{g.username}成功绑定了资产表{table.name}的IP列{ip_column}")

    return jsonify({"code": 0, "message": "绑定成功"}), 200

@manage.route("/title", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查询资产表标题", check_ip=True)
def queryTableTitle():
    '''查询所有资产表标题'''
    args = request.args          # 获取请求参数

    title = args.get("k", None)  # 搜索关键字

    if title:  # 存在关键字搜索

        return jsonify({
            "code": 0,
            "message": [item for item in list(map(lambda x: x.decode("utf-8"), redisClient.getSetData("crm:manage:table_name"))) if title in item]  # 从redis中查询
        }), 200
    
    else:

        return jsonify({"code": 0, "message": []}), 200

@manage.route("/header", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查询资产表字段", check_ip=True)
def getTableHeader():
    '''获取表格的头部字段'''
    args = request.args                # 获取请求参数

    table_uuid = args.get("id", None)  # 表格的uuid

    if not table_uuid:
        return jsonify({"code": -1, "message": "缺少id参数"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
        return jsonify({"code": -1, "message": "资产表不存在"}), 400

    try:  # 查看表是否存在
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    result = redisClient.getData(f"crm:header:{table.table_name}")  # 从redis中查询

    if result:
        result = [MyHeader(i) for i in json.loads(result)]
    else:  # 不存在,查询数据库

        try:  # 查询数据,按order升序
            result = db_session.query(Header).filter(Header.table_name == table.table_name).order_by(Header.order.asc()).all()
        finally:
            db_session.close()

        if result:
            _h = [
                {c.name: getattr(u, c.name) for c in u.__table__.columns if c.name not in ["create_user", "create_time", "update_user", "update_time"]} for u in result
            ]
            redisClient.setData(f"crm:header:{table.table_name}", json.dumps(_h))  # 写入缓存

    if not result:  # 结果为空
        return jsonify({"code": 0, "message": []}), 200

    data = []

    for item in result:

        obj = {
            "field": item.value,                 # 值
            "title": item.name,                  # 标题
            "fieldTitle": item.name,             # 标题
            "col_type": item.type,               # 列类型:1-字符串,2-下列列表
            "value_type": item.value_type,       # 值类型
            "length": item.length,               # 值长度
            "is_unique": bool(item.is_unique),   # 是否唯一
            "is_mark": bool(item.is_desence),    # 是否加密
            "is_ip": bool(item.is_ip),           # 是否ip
            "must_input": bool(item.must_input)  # 是否必填
        }

        if item.type == 2:  # 如果是下拉框

            try:
                options = db_session.query(Options.option_name, Options.option_value).filter(Options.table_name == table.table_name, Options.header_value == item.value).all()
            finally:
                db_session.close()

            _obj = {}

            for opt in options:
                _obj[opt.option_value] = opt.option_name
            obj["option"] = _obj
        
        data.append(obj)

    return jsonify({"code": 0,"message": data}), 200

@manage.route("/<string:id>", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查看资产表详情", check_ip=True)
def queryTableByUuid(id):
    '''查看指定的资产表'''
    if not redisClient.getSet("crm:manage:table_uuid", id):  # 如果不存在
        return jsonify({"code": -1, "message": "资产表不存在"}), 400

    try:
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == id).first()  # 根据id查找资产表
    finally:
        db_session.close()

    if not table:  # 表不存在
        return jsonify({"code": -1, "message": "资产表不存在"}), 400

    args = request.args                        # 获取请求参数

    page = int(args.get("page", default=1))     # 当前页码,默认第一页
    limit = int(args.get("limit", default=6))   # 每页显示数量,默认6条

    columns = redisClient.getData(f"crm:header:{table.table_name}")  # 从redis中查询

    if columns:

        columns = [MyHeader(i) for i in json.loads(columns)]
    
    else:  # 不存在,查询数据库

        try:
            columns = db_session.query(Header.value, Header.is_desence).filter(Header.table_name == table.table_name).all()  # 获取表头信息,所有列
        finally:
            db_session.close()

    if not columns:  # 如果没有表头信息,则返回空列表
        return jsonify({"code": 0, "message": {"total": 0, "data": []}}), 200

    # 根据用户搜索关键字返回数据
    key_type = int(args.get("type", default=1)) # 字段类型:1-字符串,2-下列列表,3-日期
    key = args.get("key", None)                 # 用户查找的字段
    value = args.get("value", None)             # 用户查找的值

    if key_type == 3:
        compare = args.get("c", None)
        if not compare:
            return jsonify({"code": -1, "message": "缺少c参数"}), 400
        
    manageTable = initManageTable(table.table_name)  # 实例化已存在资产表

    if key and value:  # 如果存在关键字搜索

        try:
            if key_type == 1:
                compare = args.get("c", None)
                if compare and compare == "eq":
                    count = db_session.query(manageTable).filter(getattr(manageTable.c, key) == value).count()
                else:
                    count = db_session.query(manageTable).filter(getattr(manageTable.c, key).like(f"%{value}%")).count()
            elif key_type == 2:
                count = db_session.query(manageTable).filter(getattr(manageTable.c, key) == value).count()
            elif key_type == 3:
                if compare == "eq":
                    count = db_session.query(manageTable).filter(getattr(manageTable.c, key) == value).count()
                elif compare == "gt":
                    count = db_session.query(manageTable).filter(getattr(manageTable.c, key) > value).count()
                elif compare == "lt":
                    count = db_session.query(manageTable).filter(getattr(manageTable.c, key) < value).count()
                elif compare == "ge":
                    count = db_session.query(manageTable).filter(getattr(manageTable.c, key) >= value).count()
                elif compare == "le":
                    count = db_session.query(manageTable).filter(getattr(manageTable.c, key) <= value).count()
                elif compare == "ne":
                    count = db_session.query(manageTable).filter(getattr(manageTable.c, key) != value).count()
        finally:
            db_session.close()

        if count == 0:  # 如果没有搜索到结果,直接返回空列表
            return jsonify({"code": 0, "message": {"total": 0, "data": []}}), 200

        try:
            if key_type == 1:
                compare = args.get("c", None)
                if compare and compare == "eq":
                     result = db_session.query(manageTable).filter(getattr(manageTable.c, key) == value).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
                else:
                    result = db_session.query(manageTable).filter(getattr(manageTable.c, key).like(f"%{value}%")).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
            elif key_type == 2:
                result = db_session.query(manageTable).filter(getattr(manageTable.c, key) == value).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
            elif key_type == 3:
                if compare == "eq":
                    result = db_session.query(manageTable).filter(getattr(manageTable.c, key) == value).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
                elif compare == "gt":
                    result = db_session.query(manageTable).filter(getattr(manageTable.c, key) > value).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
                elif compare == "lt":
                    result = db_session.query(manageTable).filter(getattr(manageTable.c, key) < value).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
                elif compare == "ge":
                    result = db_session.query(manageTable).filter(getattr(manageTable.c, key) >= value).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
                elif compare == "le":
                    result = db_session.query(manageTable).filter(getattr(manageTable.c, key) <= value).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
                elif compare == "ne":
                    result = db_session.query(manageTable).filter(getattr(manageTable.c, key) != value).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
        finally:
            db_session.close()
    
    else:  # 不存在关键字搜索

        try:
            count = db_session.query(manageTable).count()
        finally:
            db_session.close()

        if count == 0:
            return jsonify({"code": 0, "message": {"total": 0, "data": []}}), 200

        try:
            result = db_session.query(manageTable).order_by(manageTable.c._id.asc()).offset((page - 1) * limit).limit(limit).all()
        finally:
            db_session.close()

    data = []

    for item in result:

        obj = {}
        obj["_id"] = item._id

        for col in columns:
            curr_value = getattr(item, col.value) if getattr(item, col.value) else ""  # 获取对应属性值
            if curr_value:
                if col.value_type == 4:
                    curr_value = formatDate(curr_value)
                elif col.value_type == 5:
                    curr_value = formatDate(curr_value, 2)
                if bool(col.is_desence):                # 数据脱敏展示
                    curr_value = undesense(curr_value)
            obj[col.value] = curr_value

        data.append(obj)

    try:
        visit_log = Log(ip=g.ip_addr, operate_type="访问资产表", operate_content=f"访问资产表({table.name})", operate_user=g.username)
        db_session.add(visit_log)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[queryTableByUuid]写入log表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()

    crmLogger.info(f"用户{g.username}访问资产表({table.name})")

    return jsonify({"code": 0, "message": {"total": count, "data": data}}), 200

@manage.route("/undesense", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查看脱敏字段", check_ip=True)
def getDesenseField():
    '''查看指定脱敏字段'''
    args = request.args

    table_uuid = args.get("table_id", None)
    row_id = args.get("row_id", None)
    key_arr = args.get("key", None)

    if not all([table_uuid, row_id, key_arr]):  # 校验参数
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    manageTable = initManageTable(table_name=table.table_name)

    try:
        row = db_session.query(manageTable).filter(manageTable.c._id == row_id).first()
    finally:
        db_session.close()

    obj = {}
    for k in key_arr.split(","):
        if hasattr(row, k):
            obj[k] = getattr(row, k)
    
    return jsonify({"code": 0, "message": obj}), 200

@manage.route("/add", methods=methods.ALL)
@verify(allow_methods=["POST"], module_name="创建资产表", check_ip=True)
def addTableData():
    '''创建资产表'''
    reqData = request.get_json()                # 获取请求参数

    if not all(key in reqData for key in ["filename", "name", "keyword", "desc"]):  # 校验参数
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400

    filename = reqData["filename"]              # excel文件名
    table_name = reqData["name"]                # 资产表标题
    table_keyword = reqData["keyword"].lower()  # 资产表别名
    table_desc = reqData["desc"]                # 资产表描述

    if not table_name or not table_keyword:
        return jsonify({"code": -1, "message": "表名或表别名不能为空"}), 400

    if table_keyword in SYSTEM_DEFAULT_TABLE:
        return jsonify({"code": -1, "message": "不能使用系统表表名"}), 400

    try:  # 判断表是否已存在,使用or查询
        is_exist_table = db_session.query(Manage).filter(or_(Manage.name == table_name, Manage.table_name == table_keyword)).first()
    finally:
        db_session.close()

    if is_exist_table:  # 如果已经存在资产表
        return jsonify({"code": -1, "message": "表名或表别名已存在"}), 400

    if filename:  # 表格导入资产表方式

        try:
            file = db_session.query(File.affix).filter(File.uuid == filename).first()  # 根据文件id查询文件
        finally:
            db_session.close()

        if not file:
            return jsonify({"code": -1, "message": "导入的表格文件不存在"}), 400

        temp_table = readExcel(os.path.join(UPLOAD_EXCEL_DIR, f"{filename}.{file.affix}"))  # 读取表格

        if temp_table is None:
            return jsonify({"code": -1, "message": "读取表格失败"}), 400

        table_headers = temp_table.columns.tolist()  # 表头字段

        if len(table_headers) == 0:
            crmLogger.error(f"[addTableData]读取表格文件{filename}失败: 原因: 表头读取为空")
            return jsonify({"code": -1, "message": "读取表格失败"}), 400

        column_header = converWords(table_headers)  # 中文拼音转换

        try:  # 批量插入表头数据
            header_data = [Header(name=k, value=v["pinyin"], table_name=table_keyword, order=v["index"], create_user=g.username) for k, v in column_header.items()]
            db_session.add_all(header_data)
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[addTableData]写入header表发生异常: {traceback.format_exc()}")
            return jsonify({"code": -1, "message": "数据库异常"}), 500
        finally:
            db_session.close()

        manageTable = generateManageTable(table_keyword, [Column(h["pinyin"], String(255)) for _, h in column_header.items()])  # 创建资产表
        
        if manageTable is None:
            return jsonify({"code": -1, "message": "创建资产表失败"}), 400

        table_uuid = getUuid()

        try:  # 写入manage表
            db_session.add(Manage(uuid=table_uuid, name=table_name, table_name=table_keyword, description=table_desc, create_user=g.username))
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[addTableData]写入manage表发生异常: {traceback.format_exc()}")
            return jsonify({"code": -1,"message": "数据库异常"}), 500
        finally:
            db_session.close()

        try:  # 批量插入数据

            with engine.begin() as conn:  # 开启事务
                insert_data = temp_table.to_dict(orient="records")  # 转换成[{k1: v1, k2: v2...}]
                for i in insert_data:
                    for c, v in column_header.items():
                        new_data = i.pop(c)  # c-中文字段名
                        if new_data:
                            if isinstance(new_data, datetime):  # 如果数据格式是时间,格式化为字符串
                                new_data = new_data.strftime("%Y-%m-%d %H:%M:%S")
                            elif isinstance(new_data, date):    # 如果数据格式是日期,格式化为字符串
                                new_data = new_data.strftime("%Y-%m-%d")
                            else:
                                new_data = str(new_data)
                        else:
                            new_data = ""
                        i.update({v["pinyin"]: new_data})   # 将中文k1更新为拼音
                stmt = insert(manageTable)
                conn.execute(stmt, insert_data)

        except:

            crmLogger.error(f"[addTableData]批量写入资产表{table_name}发生异常: {traceback.format_exc()}")
            return jsonify({"code": -1,"message": "数据库异常"}), 500
        
    else:  # 直接创建资产表,后面创建列方式

        manageTable = generateManageTable(table_keyword)

        if manageTable is None:
            return jsonify({"code": -1, "message": "创建资产表失败"}), 400

        table_uuid = getUuid()

        try:  # 写入manage表
            db_session.add(Manage(uuid=table_uuid, name=table_name, table_name=table_keyword, description=table_desc, create_user=g.username))
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[addTableData]写入manage表发生异常: {traceback.format_exc()}")
            return jsonify({"code": -1, "message": "数据库异常"}), 500
        finally:
            db_session.close()

    try:  # 写入log表
        create_log = Log(ip=g.ip_addr, operate_type="创建资产表", operate_content=f"创建资产表{table_name}", operate_user=g.username)
        db_session.add(create_log)
    except:
        db_session.rollback()
        crmLogger.error(f"[addTableData]写入log表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()

    redisClient.setSet("crm:manage:table_name", table_name)  # 将标题写入缓存
    redisClient.setSet("crm:manage:table_uuid", table_uuid)  # 将表id写入缓存
     
    crmLogger.info(f"[addTableData]用户{g.username}创建资产表{table_name}成功")  # 日志文件记录

    return jsonify({"code": 0, "message": table_uuid}), 200

@manage.route("/template", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="下载资产表模板", check_ip=True)
def downloadTableTemplate():
    '''下载资产表模板'''
    args = request.args        # 获取请求参数

    table_uuid = args.get("id", None)  # 获取表uuid

    if not table_uuid:
        return jsonify({"code": -1, "message": "缺少id参数"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
        return jsonify({"code": -1, "message": "资产表不存在"}), 400

    try:
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:  # 查询的表不存在
        return jsonify({"code": -1, "message": "资产表不存在"}), 400

    templ_file = redisClient.getData(f"crm:{table_uuid}:template")  # 如果redis有缓存,直接返回模板文件id

    if templ_file:  # 缓存存在则直接返回

        try:
            file = db_session.query(File.filename).filter(File.uuid == templ_file).first()
        finally:
            db_session.close()

        return jsonify({"code": 0, "message": {"filename": file.filename, "fileuuid": templ_file}}), 200
   
    else:  # 缓存不存在则创建模板文件

        header = redisClient.getData(f"crm:header:{table.table_name}")

        if header:
            header = [MyHeader(i) for i in json.loads(header)]
        else:
            try:  # 查询header
                header = db_session.query(Header.name, Header.type, Header.value, Header.value_type, Header.must_input, Header.order).filter(Header.table_name == table.table_name).order_by(Header.order.asc()).all()
            finally:
                db_session.close()

        if not header:
            return jsonify({"code": -1, "message": "资产表暂无字段"}), 400

        table_header = {}
        table_styles = {}

        for h in header:    
            table_header[h.name] = { 
                "index": get_column_letter(h.order + 1),  # 转换成大写字母
                "must_input": h.must_input == 1
            }

            if h.type == 2:  # 如果是下拉列表

                try:
                    opt = db_session.query(Options.option_name).filter(Options.table_name == table.table_name, Options.header_value == h.value).all()
                finally:
                    db_session.close()

                table_styles[h.name] = {
                    "index": get_column_letter(h.order + 1),
                    "options": ",".join([o.option_name for o in opt])
                }

        fileUuid = getUuid()

        if not createExcel(TEMP_DIR, fileUuid, "导入模板", table_header, {}, table_styles, True):
            return jsonify({"code": -1, "message": "创建模板文件失败"}), 500

        redisClient.setData(f"crm:{table_uuid}:template", fileUuid)   # 写入redis

        try:  # 写入file表
            db_session.add(File(uuid=fileUuid, filename=f"{table.name}导入模板.xlsx", affix="xlsx", filepath=0))
            db_session.commit()
        except:
            db_session.rollback()
            return jsonify({"code": -1, "message": "数据库异常"}), 500
        finally:
            db_session.close()

        crmLogger.info(f"[downloadTableTemplate]生成资产表<{table.name}>导入模板文件完成")

        return jsonify({"code": 0, "message": {"filename": f"{table.name}导入模板.xlsx", "fileuuid": fileUuid}}), 200

@manage.route("/import", methods=methods.ALL)
@verify(allow_methods=["POST"], module_name="导入资产表数据", check_ip=True)
def importTableFromExcel():
    '''导入资产表'''
    reqData = request.get_json()  # 获取请求数据

    if not all(key in reqData for key in ["file_uuid", "table_id"]):  # 校验body参数
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400

    file_uuid = reqData["file_uuid"]    # 获取上传的文件uuid
    table_id = reqData["table_id"]      # 获取导入的资产表id

    if not all([file_uuid, table_id]):   # 必填参数值为空情况
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_id):     # 如果资产表不存在
        return jsonify({"code": -1, "message": "资产表不存在"}), 400

    try:                    # 查询资产表是否存在
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_id).first()
    finally:
        db_session.close()  # 关闭数据库连接

    if not table:           # 资产表不存在
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:                    # 查询文件是否存在
        file = db_session.query(File.affix).filter(File.uuid == file_uuid).first()
    finally:
        db_session.close()  # 关闭数据库连接
    
    if not file:            # 文件不存在
        return jsonify({"code": -1, "message": "导入的文件不存在"}), 400
    
    task_id = getUuid()

    try:                    # 写入记录表
        import_history = History(id=task_id, file_uuid=file_uuid, mode=1, status=0, table_name=table.table_name, create_user=g.username)
        db_session.add(import_history)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[importTableFromExcel]写入history表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()

    try:
        import_log = Log(ip=g.ip_addr, operate_type="导入表格", operate_content=f"模板导入资产表{table.name}", operate_user=g.username)
        db_session.add(import_log)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[importTableFromExcel]写入log表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()
    
    # 将任务信息插入到队列
    redisClient.lpush(f"crm:import:{table.table_name}", json.dumps({
        "task_id": task_id,                   # 任务id
        "file": f"{file_uuid}.{file.affix}",  # 导入的文件名
        "name": table.name,                   # 导入的表名
        "table": table.table_name,            # 导入的表别名
        "user": g.username                    # 导入的用户
    }))

    startImportTableTask(table.table_name)

    crmLogger.info(f"[importTableFromExcel]用户{g.username}创建模板导入资产表{table.name}任务{task_id}")

    return jsonify({"code": 0, "message": task_id}), 200

@manage.route("/add_or_edit", methods=methods.ALL)
@verify(allow_methods=["POST"], module_name="修改或新增资产表数据", check_ip=True)
def addOrEditTableData():
    '''修改或新增资产表数据'''
    reqData = request.get_json()  # 获取请求数据

    if "mode" not in reqData or "table_id" not in reqData:
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400

    mode = reqData["mode"]
    table_id = reqData["table_id"]

    if not mode or not table_id:
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
        
    if not redisClient.getSet("crm:manage:table_uuid", table_id):
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_id).first()
    finally:
        db_session.close()

    if not table:
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    all_header = redisClient.getData(f"crm:header:{table.table_name}")  # 从redis中获取header

    if all_header:
        all_header = [MyHeader(i) for i in json.loads(all_header)]
    else:  # 如果header为空,则从数据库中获取
        try:
            all_header = db_session.query(Header).filter(Header.table_name == table.table_name).all()
        finally:
            db_session.close()
    
    if not all_header:
        return jsonify({"code": -1, "message": "资产表字段为空"}), 400
    
    if not all(key in reqData for key in [h.value for h in all_header]):  # 校验body参数
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    manageTable = initManageTable(table.table_name)

    class MyTable(object):
        pass

    mapper(MyTable, manageTable)

    # 校验数据
    for h in all_header:

        if h.must_input == 1 and not reqData[h.value]:  # 校验必填项
            return jsonify({"code": -1, "message": f"{h.name}字段为必填项"}), 400

        if h.value_type == 2:  # 校验数据长度

            if not reqData[h.value]:
                return jsonify({"code": -1, "message": f"{h.name}字段长度应为{h.length}"}), 400
            elif reqData[h.value] and len(reqData[h.value]) != h.length:
                return jsonify({"code": -1, "message": f"{h.name}字段长度不符合要求"}), 400

        if h.type == 2:  # 校验是否从选项中取值
            if reqData[h.value]:
                try:
                    _opt = db_session.query(Options.option_name, Options.option_value).filter(Options.table_name == table.table_name, Options.header_value == h.value).all()
                finally:
                    db_session.close()

                if reqData[h.value] not in [o.option_value for o in _opt]:
                    return jsonify({"code": -1, "message": f"{h.name}字段值不在固定选项中"}), 400
            
        if h.is_unique == 1:  # 校验是否唯一         
            if reqData[h.value]:

                if mode == "add":    # 插入数据
                    try:
                        _data = db_session.query(manageTable).filter(getattr(manageTable.c, h.value) == reqData[h.value]).first()
                        if _data:
                            return jsonify({"code": -1, "message": f"{h.name}字段值重复"}), 400
                    finally:
                       db_session.close()

                elif mode == "edit": # 编辑数据

                    if "id" not in reqData:
                        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
                    try:
                        _data = db_session.query(manageTable).filter(getattr(manageTable.c, h.value) == reqData[h.value], manageTable.c._id != reqData["id"]).first()
                        if _data:
                            return jsonify({"code": -1, "message": f"{h.name}字段值重复"}), 400 
                    finally:

                        db_session.close()

    if mode == "add":  # 插入数据

        my_table = MyTable()

        try:

            for i in all_header:
                if hasattr(manageTable.c, i.value):
                    if reqData[i.value]:
                        if i.type == 2:
                            for v in _opt:
                                if v.option_value == reqData[i.value]:
                                    setattr(my_table, i.value, f"{v.option_name}")
                                    break
                        else:
                            setattr(my_table, i.value, reqData[i.value])
            
            db_session.add(my_table)
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[addOrEditTableData]写入资产表{table.table_name}发生异常: {traceback.format_exc()}")
            return jsonify({"code": -1, "message": "数据库异常"}), 500
        finally:
            db_session.close()

    elif mode == "edit":  # 编辑数据

        try:
            data = db_session.query(manageTable).filter(getattr(manageTable.c, "_id") == reqData["id"]).first()  # 根据行id筛选数据
        finally:
            db_session.close()

        if data:
            # 更新信息
            new_data = {}
            for i in all_header:
                if hasattr(data, i.value):
                    if reqData[i.value]:
                        _curr = reqData[i.value]
                        if i.type == 2:  # 下拉列表
                            for v in _opt:
                                if v.option_value == reqData[i.value]:
                                    _curr = v.option_name
                                    break
                        try:
                            if i.value_type == 4:    # 日期
                                _curr = datetime.strptime(reqData[i.value], "%Y-%m-%d")
                            if i.value_type == 5:  # 时间
                                _curr = datetime.strptime(reqData[i.value], "%Y-%m-%d %H:%M:%S")
                        except:
                            return jsonify({"code": -1, "message": f"字段{i.name}日期格式错误"}), 400
                        new_data[i.value] = _curr
                    else:
                        if i.value_type in [4, 5]:
                            new_data[i.value] = None
                        else:
                            new_data[i.value] = ""
        else:
            return jsonify({"code": -1, "message": "数据不存在"}), 400

        try:
            db_session.query(manageTable).filter(getattr(manageTable.c, "_id") == reqData["id"]).update(new_data)   
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[addOrEditTableData]更新资产表{table.name}数据失败: {traceback.format_exc()}")
            return jsonify({"code": -1, "message": "数据库异常"}), 500
        finally:
            db_session.close()

    try:  # 写入log表

        _method = "新增" if mode == "add" else "修改"

        add_or_edit_log = Log(ip=g.ip_addr, operate_type=f"{_method}数据", operate_content=f"{_method}{table.name}数据", operate_user=g.username)
        db_session.add(add_or_edit_log)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[addOrEditTableData]写入log表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()

    # 新增或者修改数据成功要删除图表缓存
    redisClient.delData(f"crm:echart:{table.table_name}")

    crmLogger.info(f"[addOrEditTableData]用户{g.username}{_method}资产表{table.name}数据成功")

    return jsonify({"code": 0, "message": f"{_method}数据成功"}), 200

@manage.route("/delete", methods=methods.ALL)
@verify(allow_methods=["POST"], module_name="删除资产表数据", check_ip=True)
def deleteTableData():
    '''删除资产表数据'''
    reqData = request.get_json()

    if not all(key in reqData for key in ["table_uuid", "data"]):  # 校验body参数
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    table_uuid = reqData["table_uuid"]
    delData = reqData["data"]

    if not all([table_uuid, delData]):  # 校验参数是否有值
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    manageTable = initManageTable(table.table_name)

    try:
        db_session.query(manageTable).filter(getattr(manageTable.c, "_id").in_(delData)).delete(synchronize_session=False)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[deleteTableData]删除资产表{table.table_name}发生异常: {traceback.format_exc()}")
        return jsonify({"code": -1, "message": "数据库异常"}), 500
    finally:
        db_session.close()

    try:
        delete_log = Log(ip=g.ip_addr, operate_type="删除数据", operate_content=f"删除{table.table_name}资产表数据", operate_user=g.username)
        db_session.add(delete_log)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[deleteTableData]写入log表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()

    # 删除数据也要清除图表缓存
    redisClient.delData(f"crm:echart:{table.table_name}")

    crmLogger.info(f"[deleteTableData]用户{g.username}删除{table.name}数据成功")

    return jsonify({"code": 0, "message": "删除成功"}), 200

@manage.route("/add_or_alter_column", methods=methods.ALL)
@verify(allow_methods=["POST"], module_name="添加或修改列", check_ip=True)
def addOrAlterTableColumn():
    '''添加或修改资产表列信息'''
    reqData = request.get_json()  # 获取请求数据

    # 校验body参数
    if not all(key in reqData for key in ["mode", "table_uuid", "col_name", "col_alias", "type", "options", "must_input", "is_desence", "is_unique", "length"]):
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400

    mode = reqData["mode"]              # 模式:add-新增,alter-修改
    table_uuid = reqData["table_uuid"]  # 资产表uuid
    col_name = reqData["col_name"]      # 列名
    col_alias = reqData["col_alias"]    # 列别名
    col_type = reqData["type"]          # 列类型: 1-字符串(VARCHAR(255)),2-固定长度字符串,3-大文本text(超过255个字符),4-日期(yyyy-mm-dd),5-时间(yyyy-mm-dd hh:mm:ss),6-下拉选项
    options = reqData["options"]        # 如果列类型是6,此选项为选项列表
    must_input = reqData["must_input"]  # 是否必填: 0-否,1-是
    is_desence = reqData["is_desence"]  # 是否脱敏: 0-否,1-是
    is_unique = reqData["is_unique"]    # 是否唯一: 0-否,1-是
    length = reqData["length"]          # 如果列类型为2,此选项为长度

    # 校验必填参数是否有值
    if not all([mode, table_uuid, col_name, col_alias, type, must_input, is_desence, is_unique]):
        return jsonify({"code": -1, "message": "请求参数不完整"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):  # 查询表uuid是否存在
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    col_type = int(col_type)
    if col_type not in [1, 2, 3, 4, 5, 6]:  # 校验列类型
        return jsonify({"code": -1, "message": "数据类型错误"}), 400
    must_input = int(must_input)
    is_desence = int(is_desence)
    is_unique = int(is_unique)
    length = int(length)
    
    if col_type == 2 and length > 50 and length < 1:
        return jsonify({"code": -1, "message": "长度范围是1~50"}), 400
    
    if col_type == 6 and len(options) == 0:
        return jsonify({"code": -1, "message": "下拉选项不能为空"}), 400
    
    try:  # 查询相应的资产表是否存在
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:  # 如果不存在资产表
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    lock = redisClient.lock(f"{table.table_name}_lock", timeout=30)
    
    if lock.acquire(blocking=False):

        try:
            if mode == "add":
                try:  # 判断要新增的列是否有重复
                    is_exist_col = db_session.query(Header).filter(Header.table_name == table.table_name, or_(Header.name == col_name, Header.value == col_alias)).first()
                finally:   
                    db_session.close()

                if is_exist_col:  # 如果已经存在新增的列名或列别名
                    return jsonify({"code": -1, "message": "该列已存在,请勿重复创建"}), 400

                add_col_type = "VARCHAR(255)"
                if col_type == 3:
                    add_col_type = "TEXT"
                elif col_type == 4:
                    add_col_type = "DATE"
                elif col_type == 5:
                    add_col_type = "DATETIME"

                # 不存在重复则创建新列
                if not addColumn(table.table_name, col_alias, add_col_type):  # 添加列
                    return jsonify({"code": -1, "message": "添加列失败"}), 400

                try:  # 查询当前order
                    curr_order = db_session.query(func.max(Header.order)).filter(Header.table_name == table.table_name).scalar()
                finally:
                    db_session.close()

                try:  # 写入header表
                    _col_type = 1
                    _value_type = col_type
                    if col_type == 6:
                        _col_type = 2
                        _value_type = 1

                    new_header = Header(type=_col_type, name=col_name, value=col_alias, value_type=_value_type, table_name=table.table_name, is_unique=int(is_unique), is_desence=int(is_desence), must_input=int(must_input), length=int(length), order=(curr_order + 1), create_user=g.username)
                    db_session.add(new_header)
                    db_session.commit()
                except:
                    db_session.rollback()
                    crmLogger.error(f"[addOrAlterTableColumn]写入header表发生异常: {traceback.format_exc()}")
                    return jsonify({"code": -1, "message": "数据库异常"}), 500
                finally:
                    db_session.close()

                if col_type == 6:  # 如果是下拉列表,则需要写入下拉列表数据
                
                    try:
                        new_options = [Options(option_name=o["name"], option_value=o["value"], header_value=col_alias, table_name=table.table_name) for o in options]
                        db_session.add_all(new_options)
                        db_session.commit()
                    except:
                        db_session.rollback()
                        crmLogger.error(f"[addOrAlterTableColumn]写入options表发生异常: {traceback.format_exc()}")
                        return jsonify({"code": -1, "message": "数据库异常"}), 500
                    finally:
                        db_session.close()

            elif mode == "alter":
            
                try:  # 先查询出当前header信息
                    curr_header = db_session.query(Header).filter(Header.table_name == table.table_name, Header.value == col_alias).first()
                finally:
                    db_session.close()

                if not curr_header:
                    crmLogger.error(f"[addOrAlterTableColumn]用户{g.username}要修改的列{col_alias}不存在")
                    return jsonify({"code": -1, "message": "要修改的列不存在"}), 400

                manageTable = initManageTable(table.table_name)   # 实例化已存在的资产表

                # 如果将列设置为必填,查询此列是否存在空值
                if curr_header.must_input != must_input and must_input == 1:
                
                    try:
                        is_exist_null = db_session.query(manageTable).filter(getattr(manageTable.c, col_alias) == None).all()
                    finally:
                        db_session.close()

                    if len(is_exist_null) > 0:  # 如果存在空值则不允许修改
                        crmLogger.error(f"[addOrAlterTableColumn]用户{g.username}将列{col_alias}设置为必填: 此列存在空值")
                        return jsonify({"code": -1, "message": "该列存在空值,不允许修改"}), 400

                # 如果将列值设置为唯一,查询此列是否存在重复值
                if curr_header.is_unique != is_unique and is_unique == 1:
                
                    try:
                        is_exist_duplicate = db_session.query(getattr(manageTable.c, col_alias), func.count(getattr(manageTable.c, col_alias)).label("count")).group_by(getattr(manageTable.c, col_alias)).having(func.count(getattr(manageTable.c, col_alias)) > 1).all()
                    finally:   
                        db_session.close()

                    if len(is_exist_duplicate) > 0:
                        crmLogger.error(f"[addOrAlterTableColumn]用户{g.username}将列{col_alias}设置为唯一: 此列存在重复值")
                        return jsonify({"code": -1, "message": "该列存在重复值,不允许修改"}), 400

                # 如果设置列值长度为固定长度,查询此列数据长度是否满足要求
                if curr_header.value_type != col_type and col_type == 2:
                
                    try:
                        is_exist_unlength = db_session.query(func.count()).filter(func.length(getattr(manageTable.c, col_alias)) != length).scalar()
                    finally:
                        db_session.close()

                    if is_exist_unlength > 0:  # 如果存在不满足长度的值则不允许修改
                        crmLogger.error(f"[addOrAlterTableColumn]用户{g.username}将列{col_alias}设置为固定长度: 此列存在不满足长度的值")
                        return jsonify({"code": -1, "message": "存在不满足长度的值,不允许修改"}), 400

                if curr_header.value_type != col_type and col_type in [4, 5]:

                    try:  # 首先查询此列中所有不为空的数据
                        query_not_null_sql = f"SELECT {col_alias} FROM {table.table_name} WHERE {col_alias} IS NOT NULL"
                        query_not_null_data = db_session.execute(query_not_null_sql).fetchall()
                    finally:
                         db_session.close()

                    # 校验列数据是否可以被转换为时间
                    if col_type == 4:
                    
                        date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')

                        for row in query_not_null_data:
                            if not bool(date_pattern.match(row[0])):
                                crmLogger.error(f"[addOrAlterTableColumn]用户{g.username}将列{col_alias}设置为日期: 此列存在日期格式错误")
                                crmLogger.debug(f"[addOrAlterTableColumn]错误数据: {row}")
                                return jsonify({"code": -1, "message": "存在日期格式错误,不允许修改"}), 400

                    elif col_type == 5:
                    
                        datetime_pattern = re.compile(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$')

                        for row in query_not_null_data:
                            if not bool(datetime_pattern.match(row[0])):
                                crmLogger.error(f"[addOrAlterTableColumn]用户{g.username}将列{col_alias}设置为时间: 此列存在时间格式错误")
                                crmLogger.debug(f"[addOrAlterTableColumn]错误数据: {row}")
                                return jsonify({"code": -1, "message": "存在时间格式错误,不允许修改"}), 400

                # 如果修改列是从下拉列表取值
                if curr_header.value_type != col_type and col_type == 6:
                
                    new_option_name = [o["name"] for o in options]

                    try: # 存在则判断数据是否取值与下拉列表中
                        is_all_in_options = db_session.query(getattr(manageTable.c, col_alias)).filter(getattr(manageTable.c, col_alias).notin_(new_option_name)).first()
                    finally:
                        db_session.close()

                    if is_all_in_options:
                        crmLogger.error(f"[addOrAlterTableColumn]用户{g.username}将列{col_alias}设置为下拉列表: 此列存在数据不在下拉列表中")
                        return jsonify({"code": -1, "message": "存在部分数据不在下拉列表中,不允许修改"}), 400

                    try:  # 如果type是下拉列表则更新option表             
                        db_session.query(Options).filter(Options.header_value == col_alias, Options.table_name == table.table_name).delete()
                        db_session.commit()
                    except:            
                        db_session.rollback()
                        crmLogger.error(f"[addOrAlterTableColumn]删除options表发生异常: {traceback.format_exc()}")
                        return jsonify({"code": -1, "message": "数据库异常"}), 500
                    finally:  
                        db_session.close()

                    try:  # 重新添加选项值
                        new_options = [Options(option_name=o["name"], option_value=o["value"], header_value=col_alias, table_name=table.table_name) for o in options]
                        db_session.add_all(new_options)
                        db_session.commit()
                    except:
                        db_session.rollback()
                        crmLogger.error(f"[addOrAlterTableColumn]写入options表发生异常: {traceback.format_exc()}")
                        return jsonify({"code": -1, "message": "数据库异常"}), 500
                    finally:
                        db_session.close()

                # 校验通过后更新值或者修改列属性
                new_col_type = "VARCHAR(255)"
                if curr_header.value_type != col_type:
                    if col_type == 3:
                        new_col_type = "TEXT"
                    elif col_type == 4:
                        new_col_type = "DATE"
                    elif col_type == 5:
                        new_col_type = "DATETIME"

                if not alterColumn(table.table_name, col_alias, new_col_type):  # 如果修改列失败则报错
                    return jsonify({"code": -1, "message": "修改列失败"}), 400

                try:  # 最后更新Header表中列的信息
                    db_session.query(Header).filter(Header.table_name == table.table_name, Header.value == col_alias).update({
                        "type": 1 if col_type != 6 else 2,
                        "name": col_name,
                        "value_type": 1 if col_type == 6 else col_type,
                        "is_desence": is_desence,
                        "is_unique": is_unique,
                        "must_input": must_input,
                        "length": length if col_type == 2 else 0,
                        "update_user": g.username
                    })
                    db_session.commit()
                except:
                    db_session.rollback()
                    crmLogger.error(f"[addOrAlterTableColumn]更新header表发生异常: {traceback.format_exc()}")
                    return jsonify({"code": -1, "message": "数据库异常"}), 500
                finally:
                    db_session.close()

            _method = "新增" if mode == "add" else "修改"
            
            try:  # 写入log表
                add_or_alter_log = Log(ip=g.ip_addr, operate_type=f"{_method}列", operate_content=f"{_method}资产表{table.name}列{col_alias}", operate_user=g.username)      
                db_session.add(add_or_alter_log)
                db_session.commit()       
            except:
                db_session.rollback()
                crmLogger.error(f"[addOrAlterTableColumn]写入log表发生异常: {traceback.format_exc()}")
            finally:    
                db_session.close()

            refreshMeta()

            redisClient.delData(f"crm:header:{table.table_name}")  # 从redis中删除缓存

            crmLogger.info(f"[addOrAlterTableColumn]用户{g.username}{_method}资产表{table.name}列{col_alias}成功")

            return jsonify({"code": 0, "message": f"{_method}列成功"}), 200

        finally:
            lock.release()
    else:
        return jsonify({"code": -1, "message": "当前有用户正在修改/新增列,请稍后再试"}), 400

@manage.route("/ping", methods=methods.ALL)
@verify(allow_methods=["GET", "POST"], module_name="ping探测任务", check_ip=True)
def multDetectHost():
    '''多线程ping探测'''
    if request.method == "GET":    # 获取所有任务

        args = request.args        # 获取请求参数

        table_uuid = args.get("id", None)     # 资产表id
        task_uuid = args.get("task_id", None) # 任务的uuid
        page = int(args.get("page", 1))       # 页码
        limit = int(args.get("limit", 5))     # 每页数量

        if not table_uuid:
            return jsonify({"code": -1, "message": "缺少id参数"}), 400
        
        if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
            return jsonify({"code": -1, "message": "资产表不存在"}), 400
        
        if task_uuid:  # 查询任务详情
            try:
                is_exist_task = db_session.query(Task).filter(Task.id == task_uuid).first()
            finally:
                db_session.close()

            if not is_exist_task:
                return jsonify({"code": -1, "message": "任务不存在"}), 400
            
            if is_exist_task.status in [0, 1]:
                return jsonify({"code": -1, "message": "任务未完成"}), 400
            
            pie_echart = redisClient.getData(f"crm:ping:echart:{task_uuid}")  # 查询缓存的图表

            if pie_echart:
                pie_echart = json.loads(pie_echart)
            else:
                try:
                    task_echart = db_session.query(DetectResult.status, func.count(1)).filter(DetectResult.task_id == task_uuid).group_by(DetectResult.status).all()
                finally:
                    db_session.close()

                pie_echart = [{"name": f"{'在线' if i[0] == 1 else '离线'}", "value": i[1]} for i in task_echart]

                redisClient.setData(f"crm:ping:echart:{task_uuid}", json.dumps(pie_echart))  # 写入缓存

            try:
                count = db_session.query(DetectResult).filter(DetectResult.task_id == task_uuid).count()
            finally:
                db_session.close()

            if count == 0:
                return jsonify({"code": 0, "message": {"total": 0, "data": []}}), 200
            
            try:
                task_result = db_session.query(DetectResult).filter(DetectResult.task_id == task_uuid).order_by(DetectResult.id.asc()).offset((page - 1) * limit).limit(limit).all()
            finally:
                db_session.close()

            crmLogger.info(f"[multDetectHost]用户{g.username}查看ping探测任务{task_uuid}结果详情")

            return jsonify({
                "code": 0, 
                "message": {
                    "total": count,
                    "data": [{"id": t.id, "ip": t.ip, "status": t.status, "reason": t.reason} for t in task_result],
                    "echart": pie_echart  # 饼图显示情况
                }
            }), 200

        else:

            try:
                table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
            finally:
                db_session.close()

            if not table:
                return jsonify({"code": -1, "message": "资产表不存在"}), 400

            try:
                count = db_session.query(Task).filter(Task.table_name == table.table_name, Task.create_user == g.username).count()
            finally:
                db_session.close()

            if count == 0:
                return jsonify({"code": 0, "message": {"total": 0, "data": []}}), 200

            try:
                task_list = db_session.query(Task).filter(Task.table_name == table.table_name, Task.create_user == g.username).order_by(Task.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
            finally:
                db_session.close()

            crmLogger.info(f"[multDetectHost]用户{g.username}查看资产表{table.name}的ping探测任务列表")

            return jsonify({
                "code": 0, 
                "message": {
                    "total": count,
                    "data": [{"id": t.id, "name": t.name, "create_time": formatDate(t.create_time, 2), "status": t.status} for t in task_list]
                }
            }), 200

    elif request.method == "POST":  # 创建任务

        reqData = request.get_json()  # 用户的请求body数据

        if not all(key in reqData for key in ["id", "name", "column"]):
            return jsonify({"code": -1, "message": "缺少参数"}), 400

        table_uuid = reqData["id"]
        task_name = reqData["name"] 
        ip_column = reqData["column"]  # 用户选择的IP列名

        if not all([table_uuid, task_name, ip_column]):
            return jsonify({"code": -1, "message": "缺少参数"}), 400
        
        if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
            return jsonify({"code": -1, "message": "资产表不存在"}), 400
        
        try:
            table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
        finally:
            db_session.close()
        
        if not table:
            return jsonify({"code": -1, "message": "资产表不存在"}), 400

        task_id = getUuid()  # 生成任务uuid

        try:   # 插入数据库
            task_data = Task(id=task_id, name=task_name, keyword=ip_column, table_name=table.table_name, status=0, create_user=g.username)
            db_session.add(task_data)
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[multDetectHost]写入task表发生异常: {traceback.format_exc()}")
            return jsonify({"code": -1, "message": "数据库异常"}), 500
        finally:
            db_session.close()

        try:
            add_task_log = Log(ip=g.ip_addr, operate_type="创建Ping任务", operate_content=f"创建ping任务,任务id: {task_id}", operate_user=g.username)
            db_session.add(add_task_log)
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[multDetectHost]写入log表发生异常: {traceback.format_exc()}")
        finally:
            db_session.close()

        # 创建任务,存入redis队列
        redisClient.lpush("crm:task:ping", json.dumps({
            "task_id": task_id,
            "table": table.table_name,
            "keyword": ip_column,
            "user": g.username
        }))

        startPingTask()

        crmLogger.info(f"[multDetectHost]用户{g.username}成功创建ping任务,任务id: {task_id}")

        return jsonify({"code": 0, "message": task_id}), 200

@manage.route("/notify", methods=methods.ALL)
@verify(allow_methods=["GET", "POST"], module_name="到期通知", check_ip=True)
def notifyExpireData():
    '''到期通知'''
    if request.method == "GET":     # 获取所有通知

        args = request.args        # 获取请求参数

        table_uuid = args.get("id", None)  # 资产表id
        page = int(args.get("page", 1))    # 页码
        limit = int(args.get("limit", 5))  # 每页数量
 
        if not table_uuid:
            return jsonify({"code": -1, "message": "缺少id参数"}), 400

        if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
            return jsonify({"code": -1, "message": "资产表不存在"}), 400

        try:
            table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
        finally:
            db_session.close()

        if not table:
            return jsonify({"code": -1, "message": "资产表不存在"}), 400
        
        try:
            count = db_session.query(Notify).filter(Notify.table_name == table.table_name, Notify.create_user == g.username).count()
        finally:
            db_session.close()

        if count == 0:
            return jsonify({"code": 0, "message": {"total": 0, "data":[]}}), 200
        
        try:
            data = db_session.query(Notify).filter(Notify.table_name == table.table_name, Notify.create_user == g.username).order_by(Notify.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
        finally:
            db_session.close()

        crmLogger.info(f"[notifyExpireData]用户{g.username}查询了{table.name}到期通知信息")

        return jsonify({
            "code": 0,
            "message": {
                "total": count,
                "data": [{"id": i.id, "name": i.name, "keyword": i.keyword, "status": i.status, "create_time": formatDate(i.create_time, 2)} for i in data]
            }
        }), 200
    
    elif request.method == "POST":  # 创建通知任务

        reqData = request.get_json()    # 获取请求数据
        
        operate = reqData["operate"]    # 任务操作

        if not operate:
            return jsonify({"code": -1, "message": "请求参数不完整"}), 400
        
        if operate == "add":     # 创建任务

            table_uuid = reqData["id"]        # 资产表id
            name = reqData["name"]            # 任务名
            date_keyword = reqData["keyword"] # 日期字段 

            if not all([table_uuid, name, date_keyword]):
                return jsonify({"code": -1, "message": "请求参数不完整"}), 400
            
            if not redisClient.getSet("crm:manage:table_uuid", table_uuid):  # 判断资产表是否存在
                return jsonify({"code": -1, "message": "资产表不存在"}), 400
            
            try:
                table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
            finally:
                db_session.close()

            if not table:
                return jsonify({"code": -1, "message": "资产表不存在"}), 400
            
            try:  # 查询是否有存在已有任务
                is_exist_notify = db_session.query(Notify).filter(Notify.table_name == table.table_name, Notify.create_user == g.username, Notify.keyword == date_keyword).first()
            finally:
                db_session.close()

            if is_exist_notify:
                return jsonify({"code": -1, "message": "已存在相同任务,请勿重复创建"}), 400
            
            # 创建定时任务
            task_id = getUuid()
            
            try:
                job.setJob(id=task_id, job_time="00:05:00", func="app.src.task:notifyTask", args=[task_id, table.name, table.table_name, date_keyword]) 
            except:
                crmLogger.error(f"[notifyExpireData]创建定时任务{task_id}发生异常: {traceback.format_exc()}")
                return jsonify({"code": -1, "message": "定时任务创建失败"}), 500

            try:
                notify_data = Notify(id=task_id, name=name, table_name=table.table_name, keyword=date_keyword, create_user=g.username)
                db_session.add(notify_data)
                db_session.commit()
            except:
                db_session.rollback()
                crmLogger.error(f"[notifyExpireData]写入notify表发生异常: {traceback.format_exc()}")
                return jsonify({"code": -1, "message": "数据库异常"}), 500
            finally:
                db_session.close()

            try:  # 写入log表
                add_task_log = Log(ip=g.ip_addr, operate_type="过期通知", operate_content=f"创建定时任务{name}", operate_user=g.username)
                db_session.add(add_task_log)
                db_session.commit()
            except:
                db_session.rollback()
                crmLogger.error(f"[notifyExpireData]写入log表发生异常: {traceback.format_exc()}")
            finally:
                db_session.close()

            crmLogger.info(f"[notifyExpireData]用户{g.username}创建定时任务{name}成功")

            crmLogger.debug(f"[notifyExpireData]定时任务{name}详情: task_id={task_id}, table_name={table.table_name}, keyword={date_keyword}")

            return jsonify({"code": 0, "message": task_id}), 200
            
        elif operate in ["start", "stop"]:  # 停止任务

            task_id = reqData["task_id"]    # 任务id

            if not task_id:
                return jsonify({"code": -1, "message": "请求参数不完整"}), 400
            
            try:
                curr_task = db_session.query(Notify).filter(Notify.id == task_id).first()
            finally:
                db_session.close()

            if not curr_task:
                return jsonify({"code": -1, "message": "任务不存在"}), 400
            
            if operate == "start":
                if curr_task.status == 1:
                    return jsonify({"code": -1, "message": "任务已开启"}), 400
                else:
                    try:
                        job.resumeJob(id=curr_task.id)
                    except:
                        return jsonify({"code": -1, "message": "开启任务失败"}), 500
                    
                    try:
                        db_session.query(Notify).filter(Notify.id == task_id).update({"status": 1, "update_user": g.username})
                        db_session.commit()
                    except:
                        db_session.rollback()
                        crmLogger.error(f"[notifyExpireData]更新notify表发生异常: {traceback.format_exc()}")
                        return jsonify({"code": -1, "message": "数据库异常"}), 500
                    finally:
                        db_session.close()

            if operate == "stop":
                if curr_task == 0:
                    return jsonify({"code": -1, "message": "任务已停止"}), 400
                else:
                    try:
                        job.pauseJob(id=curr_task.id)
                    except:
                        return jsonify({"code": -1, "message": "停止任务失败"}), 500

                    try:
                        db_session.query(Notify).filter(Notify.id == task_id).update({"status": 0, "update_user": g.username})
                        db_session.commit()
                    except:
                        db_session.rollback()
                        crmLogger.error(f"[notifyExpireData]更新notify表发生异常: {traceback.format_exc()}")
                        return jsonify({"code": -1, "message": "数据库异常"}), 500
                    finally:
                        db_session.close()

            try:
                startOrStop_log = Log(ip=g.ip_addr, operate_type="过期通知", operate_content=f"{'开启' if operate == 'start' else '停止'}定时任务{task_id}", operate_user=g.username)
                db_session.add(startOrStop_log)
                db_session.commit()
            except:
                db_session.rollback()
                crmLogger.error(f"[notifyExpireData]写入log表发生异常: {traceback.format_exc()}")
            finally:
                db_session.close()
            
            crmLogger.info(f"[notifyExpireData]用户{g.username}{'开启' if operate == 'start' else '停止'}定时任务{task_id}成功")

            return jsonify({"code": 0, "message": f"{'开启' if operate == 'start' else '停止'}任务成功"}), 200

@manage.route("/export", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="导出资产表", check_ip=True)
def exportTableData():
    '''导出资产表'''
    args = request.args  # 获取请求参数

    table_uuid = args.get("id", None)   # 资产表id
    password = args.get("passwd", None) # 表格密码
    filter = args.get("filter", None)   # 筛选条件

    if not table_uuid:  # 如果请求参数没有资产表id
        return jsonify({"code": -1, "message": "缺少id参数"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:  # 查询是否存在资产表
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:  # 如果资产表不存在
        return jsonify({"code": -1, "message": "资产表不存在"}), 400

    task_id = getUuid()  # 任务id

    try:
        export_history = History(id=task_id, mode=2, status=0, table_name=table.table_name, create_user=g.username)
        db_session.add(export_history)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[exportTableData]写入history表发生异常: {traceback.format_exc()}")
        return jsonify({"code": -1, "message": "数据库异常"}), 500
    finally:
        db_session.close()

    try:
        export_log = Log(ip=g.ip_addr, operate_type="导出表格", operate_content=f"导出资产表{table.name}", operate_user=g.username)
        db_session.add(export_log)
        db_session.commit()
    except:
        db_session.rollback()
        crmLogger.error(f"[exportTableData]写入log表发生异常: {traceback.format_exc()}")
    finally:
        db_session.close()

    # 将任务信息推送到队列
    redisClient.lpush(f"crm:export:{table.table_name}", json.dumps({
        "task_id": task_id,
        "table": table.table_name,
        "name": table.name,
        "filter": filter,
        "user": g.username,
        "password": password
    }))

    startExportTableTask(table.table_name)

    crmLogger.info(f"[exportTableData]用户{g.username}创建了导出资产表{table.name}任务{task_id}")

    return jsonify({"code": 0, "message": task_id}), 200

@manage.route("/process/<string:task_id>", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查询任务进度", check_ip=True)
def progress(task_id):
    '''查询任务进度'''
    # sse推送进度
    def event_stream():
        while True:
            time.sleep(0.5)
            data = redisClient.getData(f"crm:task:{task_id}")  # 从redis中读取进度
            if not data:
                yield "data: {\"error\":\"\", \"speed\": 0}\n\n"
                continue
            yield f"data: {data}\n\n"
    return Response(event_stream(), mimetype="text/event-stream")

@manage.route("/history", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="查询历史记录", check_ip=True)
def queryHistory():
    '''查询历史记录'''
    args = request.args   # 获取请求参数

    table_uuid = args.get("id", None)  # 资产表id
    mode = int(args.get("type", 1))    # 类型:1-导入,2-导出
    page = int(args.get("page", 1))
    limit = int(args.get("limit", 5))

    if not table_uuid:
        return jsonify({"code": -1, "message": "缺少id参数"}), 400
    
    if mode not in [1, 2]:
        return jsonify({"code": -1, "message": "错误的mode参数"}), 400
    
    if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    try:
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
        
    try:
        count = db_session.query(History).filter(History.table_name == table.table_name, History.mode == mode).count()
    finally:
        db_session.close()

    if count == 0:
        return jsonify({"code": 0, "message": {"total": 0, "data": []}}), 200
    
    try:
        result = db_session.query(History).filter(History.table_name == table.table_name, History.mode == mode).order_by(History.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    finally:
        db_session.close()

    crmLogger.info(f"[queryHistory]用户{g.username}查询了资产表{table.name}的{'导入' if mode == 1 else '导出'}历史记录")

    return jsonify({"code": 0, "message": {"total": count, "data": [{"id": u.id, "file": u.file_uuid, "status": u.status, "create_user": u.create_user, "create_time": formatDate(u.create_time, 2), "error": u.err_file} for u in result]}}), 200

@manage.route("/rule", methods=methods.ALL)
@verify(allow_methods=["GET", "POST"], module_name="获取或创建图表规则", check_ip=True)
def setEchartRule():
    '''查看或创建图表规则'''
    if request.method == "GET":    # 查询图表规则

        args = request.args        # 获取请求参数

        table_uuid = args.get("id", None)  # 获取表id
        
        if not table_uuid:
            return jsonify({"code": -1, "message": "缺少id参数"}), 400
        
        if not redisClient.getSet("crm:manage:table_uuid", table_uuid):
            return jsonify({"code": -1, "message": "资产表不存在"}), 400
        
        try:           # 查询资产表表是否存在
            table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
        finally:
            db_session.close()

        if not table:  # 资产表不存在
            return jsonify({"code": -1, "message": "资产表不存在"}), 400
        
        rules = redisClient.getData(f"crm:rule:{table.table_name}")  # 获取redis缓存中的规则

        if rules:
            rules = [MyHeader(i) for i in json.loads(rules)]
        else:
            try:  # 缓存中没有则查询数据库中规则,升序
                rules = db_session.query(Echart).filter(Echart.table_name == table.table_name).order_by(Echart.id.asc()).all()
            finally:
                db_session.close()

            if rules:
                _r = [
                    {c.name: getattr(u, c.name) for c in u.__table__.columns} for u in rules
                ]
                redisClient.setData(f"crm:rule:{table.table_name}", json.dumps(_r))  # 写入缓存

        if not rules:  # 查无规则
            return jsonify({"code": 0, "message": []}), 200
        
        crmLogger.info(f"[setEchartRule]用户{g.username}查询了资产表{table.name}的图表规则")

        return jsonify({
            "code": 0,
            "message": [{"id": rule.id, "name": rule.name, "type": rule.type, "keyword": rule.keyword, "date_keyword": rule.date_keyword} for rule in rules]
        }), 200

    elif request.method == "POST":  # 创建图表规则

        reqData = request.get_json()  # 获取请求数据

        if not all(key in reqData for key in ["table_uuid", "rules"]):  # 校验body参数
            return jsonify({"code": -1, "message": "请求参数不完整"}), 400

        table_uuid = reqData["table_uuid"]  # 资产表的uuid
        rules = reqData["rules"]            # 资产表的图表规则组

        if not table_uuid or not rules or len(rules) == 0:
            return jsonify({"code": -1, "message": "请求参数不完整"}), 400
        
        if not redisClient.getSet("crm:manage:table_uuid", table_uuid):  # 资产表不存在
            return jsonify({"code": -1, "message": "资产表不存在"}), 400
        
        try:
            table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
        finally:
            db_session.close()

        if not table:
            return jsonify({"code": -1, "message": "资产表不存在"}), 400
        
        try:      # 查询资产表是否已有规则
            has_rules = db_session.query(Echart).filter(Echart.table_name == table.table_name).count()
        finally:
            db_session.close()

        if has_rules > 0:
            try:  # 有则先删除所有规则
                db_session.query(Echart).filter(Echart.table_name == table.table_name).delete()
                db_session.commit()
            except:
                db_session.rollback()
                crmLogger.error(f"[setEchartRule]删除echart表发生异常: {traceback.format_exc()}")
                return jsonify({"code": -1, "message": "数据库异常"}), 500
            finally:
                db_session.close()

        try:  # 再创建规则
            new_rules = [Echart(table_name=table.table_name, name=r["name"], type=r["type"], keyword=r["keyword"], date_keyword=r["date_keyword"]) for r in rules if r["name"] and r["type"] and r["keyword"]]
            db_session.add_all(new_rules)
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[setEchartRule]写入echart表发生异常: {traceback.format_exc()}")
            return jsonify({"code": -1, "message": "数据库异常"}), 500
        finally:
            db_session.close()

        try:
            add_log = Log(ip=g.ip_addr, operate_type="创建图表规则", operate_content=f"用户创建了{table.name}资产表图表规则", operate_user=g.username)
            db_session.add(add_log)
            db_session.commit()
        except:
            db_session.rollback()
            crmLogger.error(f"[setEchartRule]写入log表发生异常: {traceback.format_exc()}")
        finally:
            db_session.close()

        redisClient.delData(f"crm:rule:{table.table_name}")   # 删除缓存规则
        redisClient.delData(f"crm:echart:{table.table_name}") # 删除缓存图表

        crmLogger.info(f"[setEchartRule]用户{g.username}创建了资产表{table.name}的图表规则")

        return jsonify({"code": 0, "message": "规则创建成功"}), 200

@manage.route("/echart", methods=methods.ALL)
@verify(allow_methods=["GET"], module_name="获取图表信息", check_ip=True)
def getEchart():
    '''获取echart数据'''
    args = request.args        # 获取请求参数

    table_uuid = args.get("id", None)  # 获取表id

    if not table_uuid:  # 参数不存在
        return jsonify({"code": -1, "message": "缺少id参数"}), 400
    
    try:  # 查询表是否存在
        table = db_session.query(Manage.name, Manage.table_name).filter(Manage.uuid == table_uuid).first()
    finally:
        db_session.close()

    if not table:
        return jsonify({"code": -1, "message": "资产表不存在"}), 400
    
    result = redisClient.getData(f"crm:echart:{table.table_name}")  # 从缓存中查询

    if result:
        result = json.loads(result)
    else:
        rules = redisClient.getData(f"crm:rule:{table.table_name}") # 从缓存中查询图表规则

        if rules:
            rules = [MyHeader(i) for i in json.loads(rules)]
        else:
            try:  # 缓存不存在查询规则
                rules = db_session.query(Echart).filter(Echart.table_name == table.table_name).order_by(Echart.id.asc()).all()
            finally:
                db_session.close()

        if not rules:
            return jsonify({"code": 0, "message": [] }), 200

        result = []

        manageTable = initManageTable(table.table_name)  # 实例化已存在资产表

        for rule in rules:
            if rule.type == 1:  # 饼图

                try:
                    pie_result = db_session.query(getattr(manageTable.c, rule.keyword), func.count(1)).group_by(getattr(manageTable.c, rule.keyword)).all()
                finally:
                    db_session.close()

                if pie_result:
                    data = []
                    for i in pie_result:
                        data.append({"value": i[1], "name": i[0]})
                    result.append({
                        "title": {
                            "text": rule.name,
                            "left": "center"
                        },
                        "tooltip": {
                            "trigger": "item"
                        },
                        "legend": {
                            "orient": "vertical",
                            "left": "left"
                        },
                        "toolbox": {
                            "feature": {
                                "saveAsImage": {}
                            }
                        },
                        "series": [
                            {
                                "name": rule.name,
                                "type": "pie",
                                "radius": "50%",
                                "data": data,
                                "emphasis": {
                                    "itemStyle": {
                                        "shadowBlur": 10,
                                        "shadowOffsetX": 0,
                                        "shadowColor": "rgba(0, 0, 0, 0.5)"
                                    }
                                }
                            }
                        ]
                    })

            elif rule.type == 2:  # 柱形图

                try:
                    bar_result = db_session.query(getattr(manageTable.c, rule.keyword), func.count(1)).group_by(getattr(manageTable.c, rule.keyword)).all()
                finally:
                    db_session.close()

                if bar_result:
                    data_1 = []
                    data_2 = []
                    for i in pie_result:
                        data_1.append(i[0])
                        data_2.append(i[1])
                    result.append({
                        "title": {
                            "text": rule.name
                        },
                        "xAxis": {
                            "type": "category",
                            "data": data_1
                        },
                        "yAxis": {
                            "type": "value"
                        },
                        "toolbox": {
                            "feature": {
                                "saveAsImage": {}
                            }
                        },
                        "series": [
                            {
                                "data": data_2,
                                "type": "bar"
                            }
                        ]
                    })

            elif rule.type == 3:  # 折线图

                # key,   date,  count
                # 测试1  2021-01-01, 1
                # 测试2  2021-01-02, 2
                # 测试1  2021-01-02, 3
                try:  # 根据日期升序, 时间转换成date格式
                    line_result = db_session.query(getattr(manageTable.c, rule.keyword), func.date(getattr(manageTable.c, rule.date_keyword)), func.count(1)).group_by(getattr(manageTable.c, rule.keyword), func.date(getattr(manageTable.c, rule.date_keyword))).order_by(func.date(getattr(manageTable.c, rule.date_keyword)).asc()).all()
                finally:
                    db_session.close()
                
                legend_set = []
                date_set = []
                for l in line_result:
                    if l[0] not in legend_set:
                        legend_set.push(l[0])
                    if l[1] not in date_set:
                        date_set.push(l[1])

                data = {}
                for i in legend_set:
                    data[i] = {"name": i, "type": "line", "stack": "total", "data":["null" for _ in date_set]}  # 根据日期初始化都为null

                for m in line_result:
                    data[m[0]]["data"][date_set.index(m[1])] = m[2]

                result.append({
                    "title": {
                        "text": rule.name
                    },
                    "tooltip": {
                        "trigger": "axis"
                    },
                    "legend": {
                        "data": legend_set
                    },
                    "grid": {
                        "left": "3%",
                        "right": "4%",
                        "bottom": "3%",
                        "containLabel": True
                    },
                    "xAxis": {
                        "type": "category",
                        "boundaryGap": False,
                        "data": [formatDate(d) for d in date_set]
                    },
                    "yAxis": {
                        "type": "value"
                    },
                    "toolbox": {
                        "feature": {
                            "saveAsImage": {}
                        }
                    },
                    "series": [v for v in data.values()]
                })

        redisClient.setData(f"crm:echart:{table.table_name}", json.dumps(result))  # 将结果写入缓存

    crmLogger.info(f"[getEchart]用户{g.username}查询了{table.name}的图表数据")

    return jsonify({"code": 0, "message": result}), 200
